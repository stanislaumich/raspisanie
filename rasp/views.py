import time

from django.contrib import messages
from django.core.paginator import Paginator
from django.forms import SelectDateWidget, ModelForm, forms, IntegerField
from django.forms.utils import ErrorList
from django.shortcuts import render
from django.http import HttpResponseRedirect, HttpResponseNotFound, request, response, HttpResponse
from django.urls import reverse_lazy, reverse
from openpyxl.reader.excel import load_workbook

from alert import utils
import xlsxwriter
from aud.models import Aud
from grpp.models import Grp
from para.models import Para
from person.models import Person
from predmet.models import Predmet
from raspisanie.settings import MEDIA_ROOT, MEDIA_URL
import io
from django.http import FileResponse
from .models import Rasp, Reserv, XlsRaspPers
from datetime import datetime, timedelta, date
from django.views.generic import DetailView, UpdateView, DeleteView, CreateView
import locale

locale.setlocale(locale.LC_ALL, "")


def getuser(request):
    return request.session.get('userid', 0)


def datefromiso(year, week, day):
    return datetime.strptime("%d%02d%d" % (year, week, day), "%Y%W%w")


def page_not_found_view(request, exception):
    return render(request, 'rasp/404.html', status=404)


def indexRasp(request):
    r = Rasp.objects.order_by("id")
    paginator = Paginator(r, 12)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)
    return render(request, "rasp/indexRasp.html", context={"r": r, "page_obj": page_obj})


def detailRasp(request, id):
    t = id
    r = Rasp.objects.get(id=t)
    return render(request, "rasp/detailRasp.html", context={"r": r})


# ------------------------------------------


class EditRasp(UpdateView):
    model = Rasp
    queryset = Rasp.objects.all()
    fields = ['name', 'idgrp', 'idpers', 'idaud', 'idpredmet', 'idpara', 'dt']
    template_name = "rasp/editRasp.html"
    success_url = reverse_lazy('home')
    extra_context = {'zid': 0}


class AddRasp(CreateView):
    model = Rasp
    queryset = Rasp.objects.all()
    template_name = "rasp/editRasp.html"
    fields = ('name', 'idgrp', 'idpers', 'idaud', 'idpredmet', 'idpara', 'dt')
    widgets = {'dt': SelectDateWidget(years=range(2024, 2100))}
    success_url = reverse_lazy('home')
    extra_context = {'zid': 0}

    def get_initial(self):
        initial = super(AddRasp, self).get_initial()
        initial['dt'] = self.request.GET.get('dt')
        initial['idpara'] = Para.objects.get(id=self.request.GET.get('np'))
        if self.request.GET.get('idpers'):
            initial['idpers'] = Person.objects.get(id=self.request.GET.get('idpers'))
        if self.request.GET.get('idpredmet'):
            initial['idpredmet'] = Predmet.objects.get(id=self.request.GET.get('idpredmet'))
        if self.request.GET.get('idaud'):
            initial['idaud'] = Aud.objects.get(id=self.request.GET.get('idaud'))
        if self.request.GET.get('idgrp'):
            initial['idgrp'] = Grp.objects.get(id=self.request.GET.get('idgrp'))
        return initial

    def form_valid(self, form):
        f = super(AddRasp, self)
        # print(f)
        utils.send(
            toid=Person.objects.get(pk=form.instance.idpers.pk),
            fromid=Person.objects.get(pk=self.request.session.get('userid')),
            warn=0,
            short='Добавлено  расписание')
        form.save()
        return super().form_valid(form)


class DelRasp(DeleteView):
    # form_class = EditRasp
    model = Rasp
    queryset = Rasp.objects.all()
    fields = ['name', 'idgrp', 'idpers', 'idaud', 'idpredmet', 'idpara', 'dt']
    template_name = "rasp/delRasp.html"
    success_url = reverse_lazy('home')
    extra_context = {'zid': 0}


# ------------------------------------------
class Clone(forms.Form):
    n = IntegerField()

    def __init__(
            self,
            data=None,
            files=None,
            auto_id="id_%s",
            prefix=None,
            initial=None,
            error_class=ErrorList,
            label_suffix=None,
            empty_permitted=False,
            field_order=None,
            use_required_attribute=None,
            renderer=None,
    ):
        super().__init__(data, files, auto_id, prefix, initial, error_class, label_suffix, empty_permitted, field_order,
                         use_required_attribute, renderer)
        # self.n = 1

    class Meta:
        fields = ('n',)


def clonerasp(request, id):
    form = Clone(request.POST)
    if request.method == "POST":
        if form.is_valid():
            r = Rasp.objects.get(pk=id)
            r.name = 'Занятие 1'
            r.save()
            n = form.cleaned_data['n']
            c = 0
            print(n)
            for i in range(n):
                p = Rasp()
                p.idpers = r.idpers
                p.idpara = r.idpara
                p.idaud = r.idaud
                p.idgrp = r.idgrp
                p.idpredmet = r.idpredmet
                p.name = 'Занятие ' + str(c + 2)
                p.dt = r.dt + timedelta(7 * (c + 1))
                try:
                    p.save()
                    messages.success(request, f"Запись продублирована")
                    c = c + 1
                except:
                    messages.error(request, f"Ошибка записи {p.dt}, место занято")

            return HttpResponseRedirect(reverse('home'))
    else:
        form = Clone()
    return render(request, "person/clone.html")


def rspxlspers(request, id, wd):
    # дом
    fn1 = MEDIA_ROOT.__fspath__() + '/hello.xlsx'
    fn2 = MEDIA_ROOT.__fspath__() + "/"+str(getuser(request))+".xlsx"
    fn3 = MEDIA_ROOT.__fspath__() + "\\"+str(getuser(request))+".xlsx"
    # хост
    # fn1 = MEDIA_ROOT + '/hello.xlsx'
    # fn2 = MEDIA_ROOT + "/"+str(getuser(request))+".xlsx"
    # fn3 = MEDIA_ROOT + "/"+str(getuser(request))+".xlsx"
    wb = load_workbook(fn1)
    worksheet = wb.active
    # g = Rasp.objects.filter(idpers=id, dt__week=wd).order_by("dt", "idpara_id")
    k = 0
    w = []
    dtb = datefromiso(date.today().year, wd, 1).date()
    dtb = dtb + timedelta(-1 * dtb.weekday() + 0)
    for i in range(6):
        for j in range(7):
            try:
                r = Rasp.objects.get(dt=dtb, idpara=j + 1, idpers=id)
                w.append({'v': 1, 'i': r, "np": j, 'nd': i + 1})
            except:
                try:
                    r = Reserv.objects.get(dt=dtb, idpara=j + 1, idpers=id)
                    w.append({'v': 2, 'i': r, "np": j, 'nd': i + 1})
                except:
                    w.append({'v': 0, 'i': Para.objects.get(id=j + 1), "np": j + 1, 'nd': i + 1})
            k = k + 1
        dtb = dtb + timedelta(1)
    # print(w)
    k = 0
    for e in w:
        print(e['i'])
        if e['v'] == 0:
            # print('Пусто')
            r = XlsRaspPers.objects.get(idpara=e['i'].id, idpers=getuser(request), nk=1, nd=e['nd'])
            d = worksheet.cell(row=r.xlsrow, column=r.xlscol, value='')
            d = worksheet.cell(row=r.xlsrow, column=r.xlscol - 1, value=e['i'].name)
        elif e['v'] == 1:
            # print(e['i'].idgrp)
            r = XlsRaspPers.objects.get(idpara=e['i'].idpara, idpers=getuser(request), nk=1, nd=e['nd'])
            d = worksheet.cell(row=r.xlsrow, column=r.xlscol, value=e['i'].idpredmet.name + ' ' + e['i'].idgrp.name)
            d = worksheet.cell(row=r.xlsrow, column=r.xlscol-1, value=e['i'].idpara.name)
            d = worksheet.cell(row=r.xlsrow, column=r.xlscol+1, value=e['i'].idaud.name)
        else:
            # print('Резерв')
            r = XlsRaspPers.objects.get(idpara=e['i'].idpara, idpers=getuser(request), nk=1, nd=e['nd'])
            d = worksheet.cell(row=r.xlsrow, column=r.xlscol, value='Резерв')
            d = worksheet.cell(row=r.xlsrow, column=r.xlscol - 1, value=e['i'].idpara.name)
    wb.save(fn2)
    return FileResponse(open(fn3, 'rb'))

